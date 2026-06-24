#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Финальная сборка ПРОВЕРЕННОГО датасета: из confirmed.json (список верных пар,
выгруженный из tools/review.html) собирает чистую таблицу и папку — только то,
что человек подтвердил на слух. «Меньше, но качественно».

  python3 build_confirmed.py --confirmed confirmed.json
"""
import argparse
import csv
import json
import re
import shutil
import subprocess
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

OUT_XLSX = Path("dataset.xlsx")
OUT_SOUND = Path("sound")
LANG, GENDER, DIALECT = "ose", "male", "iron"


def blob_base():
    remote = subprocess.getoutput("git config --get remote.origin.url")
    mo = re.search(r"github\.com[:/]+([^/]+)/(.+?)(?:\.git)?$", remote)
    return f"https://github.com/{mo.group(1)}/{mo.group(2)}/blob/main/" if mo else ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--confirmed", default="confirmed.json")
    args = ap.parse_args()
    GH = blob_base()

    confirmed = json.loads(Path(args.confirmed).read_text(encoding="utf-8"))
    ok_paths = {c["path"] for c in confirmed}
    meta = {r["path"]: r for r in csv.DictReader(open("combined_index.csv", encoding="utf-8"))}

    src = Path("sound")
    newdir = Path("sound_verified")
    if newdir.exists():
        shutil.rmtree(newdir)
    newdir.mkdir()

    rows = []
    gi = 0
    for c in confirmed:
        old = src / c["path"]
        if not old.exists():
            continue
        gi += 1
        name = f"sound_{gi:07d}.mp3"
        shutil.copy2(old, newdir / name)
        m = meta.get(c["path"], {})
        rows.append({
            "path": name, "sentence": c["oset"], "sentence_rus": c["rus"],
            "duration": m.get("duration", ""), "source": m.get("source_pdf", ""),
        })

    # заменяем sound/ на проверенный набор
    shutil.rmtree(src)
    newdir.rename(src)

    write_xlsx(rows, GH)
    print(f"ГОТОВО: подтверждённых пар {len(rows)} -> {OUT_XLSX}, sound/ ({len(rows)} клипов)")


def write_xlsx(rows, GH):
    cols = ["path", "sentence", "sentence_rus", "comment", "sentence_domain",
            "gender", "accents variant locale", "duration", "source"]
    ru_head = ["имя файла со звуком", "предложение на осетинском",
               "перевод на русский", None, "язык", "пол", "диалект",
               "длительность звукового файла", "откуда"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"
    link_font = Font(color="0563C1", underline="single")
    for c, (rh, mh) in enumerate(zip(ru_head, cols), 1):
        if rh is not None:
            ws.cell(row=1, column=c, value=rh)
        ws.cell(row=2, column=c, value=mh)
    for i, r in enumerate(rows, start=3):
        a = ws.cell(row=i, column=1, value=r["path"])
        a.hyperlink = f"{GH}{OUT_SOUND.name}/{r['path']}"
        a.font = link_font
        ws.cell(row=i, column=2, value=r["sentence"])
        ws.cell(row=i, column=3, value=r["sentence_rus"])
        ws.cell(row=i, column=4, value="-")
        ws.cell(row=i, column=5, value=LANG)
        ws.cell(row=i, column=6, value=GENDER)
        ws.cell(row=i, column=7, value=DIALECT)
        ws.cell(row=i, column=8, value=r["duration"])
        ws.cell(row=i, column=9, value=r["source"])
    for c, w in enumerate([20, 32, 36, 14, 7, 6, 6, 19, 20], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    wb.save(OUT_XLSX)


if __name__ == "__main__":
    main()
