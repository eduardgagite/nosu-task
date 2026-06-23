#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Финальная сборка датасета из JSON-границ (авто из block_align_all или
исправленных тобой в tools/correct.html).

Для каждой фразы режет осетинский кусок [start,end] из raw/<тема>.mp3,
складывает в ОДНУ папку sound/ с уникальными именами sound_0000001.mp3…,
собирает dataset.xlsx (лист «Лист1» как у заказчика + лист «проверка»).

Запуск:
    python3 build_from_candidates.py --in tools/candidates
"""
import argparse
import csv
import json
import re
import subprocess
import urllib.parse
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_XLSX = Path("dataset.xlsx")
OUT_SOUND = Path("sound")
OUT_INDEX = Path("combined_index.csv")
LANG, GENDER, DIALECT = "ose", "male", "iron"


def topic_key(name):
    page = re.match(r"^(\d+)", name)
    page = int(page.group(1)) if page else 0
    ch = re.search(r"(?:Ч\.?\s*|_Ч_)(\d+)", name, re.I)
    ch = int(ch.group(1)) if ch else 0
    return (page, ch, name)


def blob_base():
    remote = subprocess.getoutput("git config --get remote.origin.url")
    mo = re.search(r"github\.com[:/]+([^/]+)/(.+?)(?:\.git)?$", remote)
    return (f"https://github.com/{mo.group(1)}/{mo.group(2)}/blob/main/"
            if mo else "")


def load_source_map():
    m = {}
    p = Path("dataset/_prep_report.csv")
    if p.exists():
        for r in csv.DictReader(open(p, encoding="utf-8")):
            m[r["topic"]] = r["pdf"]
    return m


def load_verify():
    """clip -> (lang, prob) из verify_clips.py — чтобы пометить русские клипы."""
    m = {}
    p = Path("_clip_verify.csv")
    if p.exists():
        for r in csv.DictReader(open(p, encoding="utf-8")):
            try:
                prob = float(r["prob"])
            except ValueError:
                prob = 0.0
            m[r["path"]] = (r.get("lang", ""), prob)
    return m


def cut(mp3, start, end, out):
    subprocess.run(["ffmpeg", "-y", "-v", "quiet", "-i", str(mp3),
                    "-ss", f"{start}", "-to", f"{end}",
                    "-c:a", "libmp3lame", "-q:a", "4", str(out)])
    r = subprocess.run(["ffprobe", "-v", "quiet", "-of", "csv=p=0",
                        "-show_entries", "format=duration", str(out)],
                       capture_output=True, text=True)
    try:
        return round(float(r.stdout.strip()), 2)
    except ValueError:
        return 0.0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="indir", default="tools/candidates")
    args = ap.parse_args()

    GH = blob_base()
    src_map = load_source_map()
    verify = load_verify()
    files = sorted(Path(args.indir).glob("*.json"),
                   key=lambda p: topic_key(p.stem))
    if OUT_SOUND.exists():
        import shutil
        shutil.rmtree(OUT_SOUND)
    OUT_SOUND.mkdir()

    records = []
    gi = 0
    for jf in files:
        obj = json.loads(jf.read_text(encoding="utf-8"))
        topic = obj["topic"]
        mp3 = Path("raw") / f"{topic}.mp3"
        if not mp3.exists():
            print(f"[пропуск] нет mp3: {topic}")
            continue
        src_pdf = src_map.get(topic, "")
        for p in obj["phrases"]:
            start, end = float(p.get("start") or 0), float(p.get("end") or 0)
            flag = p.get("flag", "-")
            if end - start < 0.2:
                # выравнивание не дало границ — строка без аудио, на проверку
                records.append({"path": "", "sentence": p["oset"],
                                "sentence_rus": p["rus"], "duration": 0.0,
                                "source": src_pdf, "topic": topic,
                                "flag": (flag if flag != "-" else
                                         "не удалось выровнять — нарезать вручную")})
                continue
            gi += 1
            name = f"sound_{gi:07d}.mp3"
            dur = cut(mp3, start, end, OUT_SOUND / name)
            f = [] if flag == "-" else [flag]
            v = verify.get(name)
            if v and v[0] == "ru" and v[1] >= 0.90:
                f.append("⚠ распознан как русский — на проверку")
            if dur < 0.3:
                f.append("очень короткий — проверить")
            elif dur > 20:
                f.append("очень длинный — проверить")
            records.append({"path": name, "sentence": p["oset"],
                            "sentence_rus": p["rus"], "duration": dur,
                            "source": src_pdf, "topic": topic,
                            "flag": "; ".join(f) if f else "-"})
        print(f"  {topic}: всего клипов {gi}")

    write_xlsx(records, GH)
    write_index(records, GH)
    flagged = sum(1 for r in records if r["flag"] != "-")
    print(f"\nГОТОВО: строк {len(records)}, аудио {gi} -> {OUT_XLSX}, {OUT_SOUND}/")
    print(f"на лист «проверка»: {flagged}")


def link(GH, name):
    return f"{GH}{OUT_SOUND.name}/{name}" if GH and name else ""


def write_xlsx(records, GH):
    cols = ["path", "sentence", "sentence_rus", "comment", "sentence_domain",
            "gender", "accents variant locale", "duration", "source"]
    ru_head = ["имя файла со звуком", "предложение на осетинском",
               "перевод на русский", None, "язык", "пол", "диалект",
               "длительность звукового файла", "откуда"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"
    link_font = Font(color="0563C1", underline="single")
    for c, (rh, mh) in enumerate(zip(ru_head, cols), 1):
        if rh is not None:
            ws.cell(row=1, column=c, value=rh)
        ws.cell(row=2, column=c, value=mh)
    for i, rec in enumerate(records, start=3):
        a = ws.cell(row=i, column=1, value=rec["path"])
        if rec["path"]:
            a.hyperlink = link(GH, rec["path"])
            a.font = link_font
        ws.cell(row=i, column=2, value=rec["sentence"])
        ws.cell(row=i, column=3, value=rec["sentence_rus"])
        ws.cell(row=i, column=4, value="-" if rec["flag"] == "-" else "на проверку")
        ws.cell(row=i, column=5, value=LANG)
        ws.cell(row=i, column=6, value=GENDER)
        ws.cell(row=i, column=7, value=DIALECT)
        ws.cell(row=i, column=8, value=rec["duration"])
        ws.cell(row=i, column=9, value=rec["source"])
    for c, w in enumerate([20, 32, 36, 22, 7, 6, 6, 19, 20], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    rv = wb.create_sheet("проверка")
    hf, hfont = PatternFill("solid", fgColor="FCE4D6"), Font(bold=True)
    heads = ["path", "ссылка (клик — послушать)", "предложение на осетинском",
             "перевод на русский", "длительность", "откуда", "что проверить"]
    for c, h in enumerate(heads, 1):
        cell = rv.cell(row=1, column=c, value=h)
        cell.fill = hf
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    r = 2
    for rec in [x for x in records if x["flag"] != "-"]:
        rv.cell(row=r, column=1, value=rec["path"])
        if rec["path"]:
            lk = rv.cell(row=r, column=2, value="послушать")
            lk.hyperlink = link(GH, rec["path"])
            lk.font = Font(color="0563C1", underline="single")
        rv.cell(row=r, column=3, value=rec["sentence"])
        rv.cell(row=r, column=4, value=rec["sentence_rus"])
        rv.cell(row=r, column=5, value=rec["duration"])
        rv.cell(row=r, column=6, value=rec["source"])
        rv.cell(row=r, column=7, value=rec["flag"])
        r += 1
    for c, w in enumerate([20, 22, 34, 32, 11, 20, 40], 1):
        rv.column_dimensions[get_column_letter(c)].width = w
    rv.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_index(records, GH):
    with open(OUT_INDEX, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["path", "url", "topic", "source_pdf", "duration",
                    "sentence", "sentence_rus", "flag"])
        for r in records:
            w.writerow([r["path"], link(GH, r["path"]), r["topic"], r["source"],
                        r["duration"], r["sentence"], r["sentence_rus"], r["flag"]])


if __name__ == "__main__":
    main()
