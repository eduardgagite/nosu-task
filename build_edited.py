#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Финальная сборка из edited.json (выгрузка редактора tools/editor.html):
режет осетинские фразы по выверенным границам в одну папку sound/ и собирает
dataset.xlsx в формате заказчика. Только то, что отмечено «оставить».

  python3 build_edited.py --edited edited.json
"""
import argparse
import csv
import json
import re
import shutil
import subprocess
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

OUT_XLSX = Path("dataset.xlsx")
OUT_SOUND = Path("sound")
LANG, GENDER, DIALECT = "ose", "male", "iron"


def tkey(name):
    pg = re.match(r"^(\d+)", name)
    ch = re.search(r"(?:Ч\.?\s*|_Ч_)(\d+)", name, re.I)
    return (int(pg.group(1)) if pg else 0, int(ch.group(1)) if ch else 0, name)


def blob_base():
    remote = subprocess.getoutput("git config --get remote.origin.url")
    mo = re.search(r"github\.com[:/]+([^/]+)/(.+?)(?:\.git)?$", remote)
    return f"https://github.com/{mo.group(1)}/{mo.group(2)}/blob/main/" if mo else ""


def load_source_map():
    m = {}
    p = Path("dataset/_prep_report.csv")
    if p.exists():
        for r in csv.DictReader(open(p, encoding="utf-8")):
            m[r["topic"]] = r["pdf"]
    return m


def ffprobe_dur(path):
    r = subprocess.run(["ffprobe", "-v", "quiet", "-of", "csv=p=0",
                        "-show_entries", "format=duration", str(path)],
                       capture_output=True, text=True)
    try:
        return round(float(r.stdout.strip()), 2)
    except ValueError:
        return 0.0


def cut_segs(mp3, segs, out):
    """Режет один или несколько кусков и склеивает В ЗАДАННОМ ПОРЯДКЕ (не по времени)."""
    if len(segs) == 1:
        s, e = segs[0]
        subprocess.run(["ffmpeg", "-y", "-v", "quiet", "-i", str(mp3),
                        "-ss", f"{s}", "-to", f"{e}",
                        "-c:a", "libmp3lame", "-q:a", "4", str(out)])
    else:
        parts = [f"[0]atrim=start={s}:end={e},asetpts=PTS-STARTPTS[a{k}]"
                 for k, (s, e) in enumerate(segs)]
        concat = "".join(f"[a{k}]" for k in range(len(segs))) + \
                 f"concat=n={len(segs)}:v=0:a=1[out]"
        flt = ";".join(parts) + ";" + concat
        subprocess.run(["ffmpeg", "-y", "-v", "quiet", "-i", str(mp3),
                        "-filter_complex", flt, "-map", "[out]",
                        "-c:a", "libmp3lame", "-q:a", "4", str(out)])
    return ffprobe_dur(out)


def item_segs(it):
    segs = it.get("segs")
    if not segs and "start" in it:
        segs = [[it["start"], it["end"]]]
    return [[float(s[0]), float(s[1])] for s in (segs or []) if float(s[1]) > float(s[0])]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--edited", default="edited.json")
    args = ap.parse_args()
    GH = blob_base()
    src_map = load_source_map()

    items = json.loads(Path(args.edited).read_text(encoding="utf-8"))
    for it in items:
        it["_segs"] = item_segs(it)
    # порядок строк: по теме, затем по самому раннему куску фразы
    items.sort(key=lambda x: (tkey(x["topic"]), min((s[0] for s in x["_segs"]), default=0)))

    if OUT_SOUND.exists():
        shutil.rmtree(OUT_SOUND)
    OUT_SOUND.mkdir()

    rows = []
    gi = 0
    for it in items:
        mp3 = Path("raw") / f"{it['topic']}.mp3"
        if not mp3.exists() or not it["_segs"]:
            continue
        gi += 1
        name = f"sound_{gi:07d}.mp3"
        dur = cut_segs(mp3, it["_segs"], OUT_SOUND / name)
        rows.append({"path": name, "sentence": it["oset"], "sentence_rus": it["rus"],
                     "duration": dur, "source": src_map.get(it["topic"], "")})

    write_xlsx(rows, GH)
    print(f"ГОТОВО: пар {len(rows)} -> {OUT_XLSX}, sound/ ({len(rows)} клипов)")


def write_xlsx(rows, GH):
    cols = ["path", "sentence", "sentence_rus", "comment", "sentence_domain",
            "gender", "accents variant locale", "duration", "source"]
    ru_head = ["имя файла со звуком", "предложение на осетинском",
               "перевод на русский", None, "язык", "пол", "диалект",
               "длительность звукового файла", "откуда"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"
    link_font = Font(color="0563C1", underline="single")
    for c, (rh, mh) in enumerate(zip(ru_head, cols), 1):
        if rh is not None:
            ws.cell(row=1, column=c, value=rh)
        ws.cell(row=2, column=c, value=mh)
    for i, r in enumerate(rows, start=3):
        a = ws.cell(row=i, column=1, value=r["path"])
        a.hyperlink = f"{GH}{OUT_SOUND.name}/{r['path']}"
        a.font = link_font
        ws.cell(row=i, column=2, value=r["sentence"])
        ws.cell(row=i, column=3, value=r["sentence_rus"])
        ws.cell(row=i, column=4, value="-")
        ws.cell(row=i, column=5, value=LANG)
        ws.cell(row=i, column=6, value=GENDER)
        ws.cell(row=i, column=7, value=DIALECT)
        ws.cell(row=i, column=8, value=r["duration"])
        ws.cell(row=i, column=9, value=r["source"])
    for c, w in enumerate([20, 32, 36, 14, 7, 6, 6, 19, 20], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    wb.save(OUT_XLSX)


if __name__ == "__main__":
    main()
