#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Validate generated Takazov dataset folders."""

import csv
import re
from pathlib import Path

from openpyxl import load_workbook


BAD_TEXT_RE = re.compile(r"[()\[\]·°\u0300\u0301]")


def read_manifest(topic):
    path = Path("work") / topic / "manifest.csv"
    if not path.exists():
        return {}
    with open(path, encoding="utf-8") as f:
        return {r["clip"]: r for r in csv.DictReader(f)}


def read_alignment(folder):
    path = folder / "alignment.csv"
    if not path.exists():
        return {}
    out = {}
    with open(path, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            out[int(r["row"])] = r
    return out


def main():
    root = Path("dataset")
    totals = {"topics": 0, "rows": 0, "audio": 0, "notes": 0}
    problems = []
    summaries = []

    for xlsx in sorted(root.glob("*/dataset.xlsx")):
        folder = xlsx.parent
        topic = folder.name
        manifest = read_manifest(topic)
        alignment = read_alignment(folder)
        wb = load_workbook(xlsx)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        totals["topics"] += 1
        note_count = 0
        audio_count = 0
        russian_assigned = []

        for idx, row in enumerate(rows, 1):
            audio, ossetian, russian, note = row
            if note:
                note_count += 1
            if audio:
                audio_count += 1
                if not (folder / audio).exists():
                    problems.append((topic, idx, "audio file missing", audio))
            else:
                problems.append((topic, idx, "empty audio", ""))
            if not ossetian:
                problems.append((topic, idx, "empty ossetian", ""))
            if not russian:
                problems.append((topic, idx, "empty russian", ""))
            for col, val in [("ossetian", ossetian or ""), ("russian", russian or "")]:
                if BAD_TEXT_RE.search(val):
                    problems.append((topic, idx, f"bad chars in {col}", val))

            ar = alignment.get(idx)
            if ar:
                for clip in (ar.get("assigned_clips") or "").split():
                    m = manifest.get(clip, {})
                    try:
                        prob = float(m.get("lang_prob") or 0)
                    except ValueError:
                        prob = 0
                    if m.get("flag_russian") == "RU?" or (m.get("lang") == "ru" and prob >= 0.85):
                        russian_assigned.append(f"{idx}:{clip}:{m.get('asr_text','')}")

        totals["rows"] += len(rows)
        totals["audio"] += audio_count
        totals["notes"] += note_count
        if russian_assigned:
            problems.append((topic, "-", "assigned RU? clips", " || ".join(russian_assigned[:10])))
        summaries.append((topic, len(rows), audio_count, note_count))

    print("topic\trows\taudio_filled\tnote_rows")
    for s in summaries:
        print(f"{s[0]}\t{s[1]}\t{s[2]}\t{s[3]}")
    print("\nTOTAL", totals)
    print("PROBLEMS", len(problems))
    for p in problems[:500]:
        print("PROBLEM\t" + "\t".join(map(str, p)))


if __name__ == "__main__":
    main()
