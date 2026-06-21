#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Собирает из всех потемных датасетов (dataset/<тема>/dataset.xlsx + sound/)
ОДИН итоговый файл по образцу заказчика («таблица проекта.xlsx»):

    dataset.xlsx
        лист «Лист1»   — основная таблица 1-в-1 как у заказчика
                         (path со ссылкой на GitHub, comment = «-»)
        лист «проверка» — строки, которые нужно сверить на слух, с флагами
    sound/             — все фрагменты в одной папке (sound_0000001.mp3 …)
    combined_index.csv — полная карта происхождения + авто-флаги (аудит)

Ссылки в колонке path ведут на файлы в GitHub-репозитории (видно из git remote).
Чтобы они открывались, репозиторий нужно запушить и сделать публичным.

Запуск:
    python3 build_combined.py
"""
import csv
import re
import shutil
import subprocess
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC_ROOT = Path("dataset")
OUT_XLSX = Path("dataset.xlsx")
OUT_SOUND = Path("sound")
OUT_INDEX = Path("combined_index.csv")
LANGID = Path("_langid_check.csv")

# фиксированные метаданные (как в образце заказчика)
LANG = "ose"          # осетинский (ISO 639)
GENDER = "male"       # мужской голос диктора
DIALECT = "iron"      # иронский диалект

RU_FLAG_PROB = 0.90   # клип уверенно распознан как русский -> вероятно лишний
SHORT_SEC = 0.6       # короче — вероятно обрезок/слово, а не фраза
LONG_SEC = 15.0       # длиннее — вероятно склейка нескольких фраз


def github_blob_base():
    """Базовый URL вида https://github.com/<owner>/<repo>/blob/<branch>/
    выводится из git remote. GitHub показывает для .mp3 плеер прямо на странице."""
    try:
        url = subprocess.run(["git", "config", "--get", "remote.origin.url"],
                             capture_output=True, text=True).stdout.strip()
        branch = subprocess.run(["git", "rev-parse", "--abbrev-ref", "HEAD"],
                                capture_output=True, text=True).stdout.strip() or "main"
        m = re.search(r"github\.com[:/]+([^/]+)/(.+?)(?:\.git)?$", url)
        if m:
            owner, repo = m.group(1), m.group(2)
            return f"https://github.com/{owner}/{repo}/blob/{branch}/"
    except Exception:
        pass
    return ""  # не удалось определить — ссылки будут локальными


GH_BASE = github_blob_base()


def topic_sort_key(name: str):
    page = re.match(r"^(\d+)", name)
    page = int(page.group(1)) if page else 0
    ch = re.search(r"(?:Ч\.?\s*|_Ч_)(\d+)", name, re.I)
    ch = int(ch.group(1)) if ch else 0
    return (page, ch, name)


def load_source_map():
    m = {}
    p = SRC_ROOT / "_prep_report.csv"
    if p.exists():
        with open(p, encoding="utf-8") as f:
            for r in csv.DictReader(f):
                m[r["topic"]] = r["pdf"]
    return m


def load_langid():
    m = {}
    if LANGID.exists():
        with open(LANGID, encoding="utf-8") as f:
            for r in csv.DictReader(f):
                try:
                    prob = float(r["prob"])
                except (ValueError, KeyError):
                    prob = 0.0
                m[r["clip"]] = (r.get("lang", ""), prob, r.get("text", ""))
    return m


def ffprobe_duration(path: Path) -> float:
    try:
        out = subprocess.run(
            ["ffprobe", "-v", "quiet", "-of", "csv=p=0",
             "-show_entries", "format=duration", str(path)],
            capture_output=True, text=True, check=True).stdout.strip()
        return round(float(out), 2)
    except Exception:
        return 0.0


def build_flag(note, langid_entry, duration, is_dup):
    """Строка авто-проверки качества (для листа «проверка» и combined_index)."""
    flags = []
    if langid_entry:
        lang, prob, _ = langid_entry
        if lang == "ru" and prob >= RU_FLAG_PROB:
            flags.append("⚠ похоже на русскую речь/диктора")
    if duration and duration < SHORT_SEC:
        flags.append("⚠ очень короткий фрагмент")
    if duration and duration > LONG_SEC:
        flags.append("⚠ длинный фрагмент (возможно склейка)")
    if is_dup:
        flags.append("⚠ текст повторяется на нескольких фрагментах")
    if note and note.strip():
        flags.append("групповая привязка — сверить на слух")
    return "; ".join(flags) if flags else "-"


def link_for(name: str) -> str:
    return f"{GH_BASE}{OUT_SOUND.name}/{name}" if GH_BASE else f"{OUT_SOUND.name}/{name}"


def main():
    source_map = load_source_map()
    langid = load_langid()
    print("GitHub base:", GH_BASE or "(не определён — ссылки локальные)")

    topics = sorted([p for p in SRC_ROOT.glob("*/dataset.xlsx")],
                    key=lambda p: topic_sort_key(p.parent.name))

    if OUT_SOUND.exists():
        shutil.rmtree(OUT_SOUND)
    OUT_SOUND.mkdir(parents=True)

    records = []
    gi = 0
    for xlsx in topics:
        folder = xlsx.parent
        topic = folder.name
        src_pdf = source_map.get(topic, "")
        wb = openpyxl.load_workbook(xlsx)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            audio, ossetian, russian, note = row
            if not audio:
                continue
            orig = folder / audio
            if not orig.exists():
                print(f"[пропуск] нет файла: {orig}")
                continue
            gi += 1
            newname = f"sound_{gi:07d}.mp3"
            shutil.copy2(orig, OUT_SOUND / newname)
            dur = ffprobe_duration(OUT_SOUND / newname)
            lid = langid.get(str(orig).replace("\\", "/"))
            records.append({
                "path": newname,
                "sentence": (ossetian or "").strip(),
                "sentence_rus": (russian or "").strip(),
                "comment": "-",                 # основная таблица — чисто, как в эталоне
                "sentence_domain": LANG,
                "gender": GENDER,
                "accents variant locale": DIALECT,
                "duration": dur,
                "source": src_pdf,
                "_topic": topic,
                "_orig": audio,
                "_note": (note or "").strip(),
                "_lid": lid,
                "_lang": lid[0] if lid else "",
                "_lang_prob": lid[1] if lid else "",
            })
        print(f"  {topic}: всего строк {gi}")

    # тексты-дубли (один и тот же текст на разных клипах — признак ошибки привязки)
    from collections import Counter
    seen = Counter(r["sentence"] for r in records if r["sentence"])
    dup_texts = {t for t, n in seen.items() if n > 1}
    for r in records:
        r["_flag"] = build_flag(r["_note"], r["_lid"], r["duration"],
                                r["sentence"] in dup_texts)

    write_xlsx(records)
    write_index(records)

    flagged = [r for r in records if r["_flag"] != "-"]
    warn = sum(1 for r in records if "⚠" in r["_flag"])
    print(f"\nГОТОВО: {len(records)} строк -> {OUT_XLSX}")
    print(f"Аудио в одной папке: {OUT_SOUND}/  ({len(records)} файлов)")
    print(f"Лист «проверка»: {len(flagged)} строк (из них ⚠ приоритетных: {warn})")
    print(f"Карта происхождения: {OUT_INDEX}")


def write_xlsx(records):
    cols = ["path", "sentence", "sentence_rus", "comment", "sentence_domain",
            "gender", "accents variant locale", "duration", "source"]
    # человеческая строка заголовка (как в эталоне; D — пусто)
    ru_head = ["имя файла со звуком", "предложение на осетинском",
               "перевод на русский", None, "язык", "пол", "диалект",
               "длительность звукового файла", "откуда"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"
    link_font = Font(color="0563C1", underline="single")

    # два ряда заголовка — без заливки и жирного, как в эталоне
    for c, (rh, mh) in enumerate(zip(ru_head, cols), 1):
        if rh is not None:
            ws.cell(row=1, column=c, value=rh)
        ws.cell(row=2, column=c, value=mh)

    for i, rec in enumerate(records, start=3):
        link = ws.cell(row=i, column=1, value=rec["path"])
        link.hyperlink = link_for(rec["path"])
        link.font = link_font
        ws.cell(row=i, column=2, value=rec["sentence"])
        ws.cell(row=i, column=3, value=rec["sentence_rus"])
        ws.cell(row=i, column=4, value=rec["comment"])
        ws.cell(row=i, column=5, value=rec["sentence_domain"])
        ws.cell(row=i, column=6, value=rec["gender"])
        ws.cell(row=i, column=7, value=rec["accents variant locale"])
        ws.cell(row=i, column=8, value=rec["duration"])
        ws.cell(row=i, column=9, value=rec["source"])

    widths = [20, 32, 36, 22, 7, 6, 6, 19, 20]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # отдельный лист с проверкой качества (то, что нужно сверить на слух)
    add_review_sheet(wb, records)
    wb.save(OUT_XLSX)


def add_review_sheet(wb, records):
    ws = wb.create_sheet("проверка")
    head_fill = PatternFill("solid", fgColor="FCE4D6")
    head_font = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")
    headers = ["path", "ссылка (клик — послушать)", "предложение на осетинском",
               "перевод на русский", "длительность", "откуда", "что проверить"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # сначала приоритетные (⚠), затем «групповая привязка»
    flagged = [r for r in records if r["_flag"] != "-"]
    flagged.sort(key=lambda r: (0 if "⚠" in r["_flag"] else 1, r["path"]))
    r = 2
    for rec in flagged:
        ws.cell(row=r, column=1, value=rec["path"])
        link = ws.cell(row=r, column=2, value="послушать")
        link.hyperlink = link_for(rec["path"])
        link.font = link_font
        ws.cell(row=r, column=3, value=rec["sentence"])
        ws.cell(row=r, column=4, value=rec["sentence_rus"])
        ws.cell(row=r, column=5, value=rec["duration"])
        ws.cell(row=r, column=6, value=rec["source"])
        ws.cell(row=r, column=7, value=rec["_flag"])
        r += 1

    for c, w in enumerate([20, 22, 34, 32, 11, 20, 46], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(1, len(flagged)+1)}"


def write_index(records):
    with open(OUT_INDEX, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["path", "url", "topic", "source_pdf", "orig_clip", "duration",
                    "sentence", "sentence_rus", "flag", "auto_lang", "auto_lang_prob"])
        for r in records:
            w.writerow([r["path"], link_for(r["path"]), r["_topic"], r["source"],
                        r["_orig"], r["duration"], r["sentence"], r["sentence_rus"],
                        r["_flag"], r["_lang"], r["_lang_prob"]])


if __name__ == "__main__":
    main()
